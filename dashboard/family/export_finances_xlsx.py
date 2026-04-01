"""
export_finances_xlsx.py — Generate a clean, formatted Finances Snapshot xlsx from finances.json

Usage:
  python export_finances_xlsx.py                     # writes to ~/Downloads/Finances_Clean.xlsx
  python export_finances_xlsx.py path/to/output.xlsx # writes to explicit path
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
except ImportError:
    print("Install openpyxl first:  pip install openpyxl")
    sys.exit(1)


def load_finances():
    fin_path = Path(__file__).resolve().parent / "finances.json"
    with open(fin_path, "r", encoding="utf-8") as f:
        return json.load(f)


def fmt_currency(ws, cell):
    cell.number_format = '$#,##0'


def add_section(ws, row, title, color, items, total, value_field="balance"):
    """Add a section header + items + total. Returns next available row."""
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(bottom=Side(style="thin", color="CCCCCC"))

    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color=color)
    row += 1

    for item in items:
        name = item.get("name", "")
        val = item.get(value_field, 0)
        ws.cell(row=row, column=2, value=name).font = Font(size=10, color="444444")
        c = ws.cell(row=row, column=3, value=val)
        fmt_currency(ws, c)
        c.font = Font(size=10, name="Consolas")
        c.alignment = Alignment(horizontal="right")
        row += 1

    # Total row
    ws.cell(row=row, column=2, value="Total").font = Font(bold=True, size=10)
    c = ws.cell(row=row, column=3, value=total)
    fmt_currency(ws, c)
    c.font = Font(bold=True, size=11, name="Consolas", color=color)
    c.alignment = Alignment(horizontal="right")
    c.border = Border(top=Side(style="thin", color="999999"))
    row += 2  # blank row after section
    return row


def build_workbook(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.sheet_properties.tabColor = "4472C4"

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18

    row = 1

    # Title
    ws.cell(row=row, column=1, value="SARTOR FAMILY — FINANCIAL SNAPSHOT").font = Font(
        bold=True, size=14, color="2F5496"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    ws.cell(row=row, column=1, value=f"As of {data['snapshot_date']}").font = Font(
        size=10, color="888888", italic=True
    )
    row += 2

    # Net Worth hero
    ws.cell(row=row, column=1, value="NET WORTH").font = Font(bold=True, size=12, color="2F5496")
    c = ws.cell(row=row, column=3, value=data["net_worth"])
    fmt_currency(ws, c)
    c.font = Font(bold=True, size=16, name="Consolas", color="22C55E")
    c.alignment = Alignment(horizontal="right")
    row += 2

    # ── ASSETS ──
    ws.cell(row=row, column=1, value="ASSETS").font = Font(bold=True, size=13, color="2F5496")
    row += 1

    # Post-Tax
    row = add_section(ws, row, "Post-Tax Investments", "22C55E",
                      data["post_tax"]["accounts"], data["post_tax"]["total"])

    # Anthropic
    ws.cell(row=row, column=1, value="Anthropic Equity (est.)").font = Font(
        bold=True, size=12, color="6366F1", italic=True
    )
    c = ws.cell(row=row, column=3, value=data["anthropic_equity"]["total"])
    fmt_currency(ws, c)
    c.font = Font(bold=True, size=11, name="Consolas", color="6366F1")
    c.alignment = Alignment(horizontal="right")
    row += 2

    # Retirement
    row = add_section(ws, row, "Retirement Accounts", "3B82F6",
                      data["retirement"]["accounts"], data["retirement"]["total"])

    # Real Estate
    row = add_section(ws, row, "Real Estate", "F59E0B",
                      data["real_estate"]["properties"], data["real_estate"]["total"],
                      value_field="value")

    # ── LIABILITIES ──
    ws.cell(row=row, column=1, value="LIABILITIES").font = Font(bold=True, size=13, color="EF4444")
    row += 1
    row = add_section(ws, row, "Debts & Obligations", "EF4444",
                      data["liabilities"]["items"], data["liabilities"]["total"])

    # ═══ CASH FLOW SHEET ═══
    ws2 = wb.create_sheet("Cash Flow")
    ws2.sheet_properties.tabColor = "22C55E"
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 16

    r = 1
    ws2.cell(row=r, column=1, value="MONTHLY CASH FLOW").font = Font(
        bold=True, size=14, color="2F5496"
    )
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2

    # Income
    ws2.cell(row=r, column=1, value="INCOME").font = Font(bold=True, size=12, color="22C55E")
    r += 1
    for item in data.get("income_breakdown", []):
        ws2.cell(row=r, column=2, value=item["name"]).font = Font(size=10, color="444444")
        c = ws2.cell(row=r, column=3, value=item["amount"])
        fmt_currency(ws2, c)
        c.font = Font(size=10, name="Consolas")
        c.alignment = Alignment(horizontal="right")
        r += 1
    ws2.cell(row=r, column=2, value="Total Income").font = Font(bold=True, size=10)
    c = ws2.cell(row=r, column=3, value=data["monthly_income"])
    fmt_currency(ws2, c)
    c.font = Font(bold=True, size=11, name="Consolas", color="22C55E")
    c.alignment = Alignment(horizontal="right")
    c.border = Border(top=Side(style="thin", color="999999"))
    r += 2

    # Expenses
    ws2.cell(row=r, column=1, value="EXPENSES").font = Font(bold=True, size=12, color="EF4444")
    r += 1
    for item in data.get("expense_breakdown", []):
        ws2.cell(row=r, column=2, value=item["name"]).font = Font(size=10, color="444444")
        c = ws2.cell(row=r, column=3, value=item["amount"])
        fmt_currency(ws2, c)
        c.font = Font(size=10, name="Consolas")
        c.alignment = Alignment(horizontal="right")
        r += 1
    ws2.cell(row=r, column=2, value="Total Expenses").font = Font(bold=True, size=10)
    c = ws2.cell(row=r, column=3, value=data["monthly_expenses"])
    fmt_currency(ws2, c)
    c.font = Font(bold=True, size=11, name="Consolas", color="EF4444")
    c.alignment = Alignment(horizontal="right")
    c.border = Border(top=Side(style="thin", color="999999"))
    r += 2

    # Gap
    gap_color = "EF4444" if data["monthly_gap"] < 0 else "22C55E"
    ws2.cell(row=r, column=1, value="MONTHLY GAP").font = Font(bold=True, size=12, color=gap_color)
    c = ws2.cell(row=r, column=3, value=data["monthly_gap"])
    fmt_currency(ws2, c)
    c.font = Font(bold=True, size=14, name="Consolas", color=gap_color)
    c.alignment = Alignment(horizontal="right")
    r += 1
    ws2.cell(row=r, column=2, value=f"Annual: ${data['monthly_gap'] * 12:,.0f}").font = Font(
        size=10, color="888888", italic=True
    )

    return wb


if __name__ == "__main__":
    data = load_finances()

    if len(sys.argv) > 1:
        out_path = Path(sys.argv[1])
    else:
        out_path = Path.home() / "Downloads" / "Finances_Clean.xlsx"

    wb = build_workbook(data)
    wb.save(str(out_path))
    print(f"Clean spreadsheet written to: {out_path}")
