"""
update_finances.py — Parse 'Finances Snapshot.xlsx' and regenerate finances.json

Usage:
  python update_finances.py                          # reads from ~/Downloads
  python update_finances.py path/to/snapshot.xlsx    # reads from explicit path

Writes finances.json to the same directory as this script.
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Install openpyxl first:  pip install openpyxl")
    sys.exit(1)


def cell(ws, coord):
    """Return numeric cell value or 0."""
    v = ws[coord].value
    return float(v) if v is not None else 0.0


def parse_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["Sheet1"]

    # Determine latest data column (rightmost with a date header in row 1)
    col = "I"  # Dec 2025 snapshot

    # ── Post-Tax Accounts ──
    post_tax_accounts = []
    for row in range(2, 11):
        name = ws[f"A{row}"].value
        val = ws[f"{col}{row}"].value
        if name and val is not None and val > 0:
            post_tax_accounts.append({"name": name, "balance": round(val)})
    post_tax_total = round(cell(ws, f"{col}11"))

    # Anthropic
    anthropic_total = round(cell(ws, f"{col}12"))

    # ── Retirement Accounts ──
    retirement_accounts = []
    retirement_rows = [13, 14, 15, 16, 17, 18, 19, 20, 21, 23, 24, 25, 26, 27]
    for row in retirement_rows:
        name = ws[f"A{row}"].value
        val = ws[f"{col}{row}"].value
        if name and val is not None and val > 0:
            retirement_accounts.append({"name": name, "balance": round(val)})
    # Sort by balance descending
    retirement_accounts.sort(key=lambda x: x["balance"], reverse=True)
    retirement_total = round(cell(ws, f"{col}28"))

    # ── Real Estate ──
    properties = []
    for row in [30, 31]:
        name = ws[f"A{row}"].value
        val = ws[f"{col}{row}"].value
        if name and val:
            properties.append({"name": name.replace("Estimated ", ""), "value": round(val)})
    re_total = sum(p["value"] for p in properties)

    # ── Liabilities ──
    liab_items = []
    liab_map = {
        33: "185 Davis Mortgage",
        34: "85 Stonebridge Mortgage",
        35: "85 Stonebridge HELOC",
        36: "Student Loan",
        44: "Car Loan",
    }
    for row, name in liab_map.items():
        val = ws[f"{col}{row}"].value
        if val and val > 0:
            liab_items.append({"name": name, "balance": round(val)})

    # Credit cards (rows 37-43)
    cc_total = 0
    for row in range(38, 44):
        val = ws[f"{col}{row}"].value
        if val is not None and val > 0:
            cc_total += val
    if cc_total > 0:
        liab_items.append({"name": "Credit Cards", "balance": round(cc_total)})

    # Sort liabilities by balance descending
    liab_items.sort(key=lambda x: x["balance"], reverse=True)
    liab_total = sum(item["balance"] for item in liab_items)

    # ── Net Worth ──
    net_worth = round(cell(ws, "J46"))

    # ── Monthly Income (column F) ──
    income_items = []
    income_map = {
        48: ("Alton salary", "F48"),
        49: ("Aneeta salary", "F49"),
        50: ("185 Davis rental", "F50"),
        51: ("Support", "F51"),
    }
    for row, (name, ref) in income_map.items():
        val = cell(ws, ref)
        if val > 0:
            income_items.append({"name": name, "amount": round(val)})
    monthly_income = sum(item["amount"] for item in income_items)

    # ── Monthly Expenses (column H — latest) ──
    expense_items = []
    expense_map = {
        55: "85 Stonebridge (PITI)",
        56: "185 Davis mortgage",
        57: "Utilities",
        58: "Yard",
        59: "185 Management",
        60: "185 Davis Condo Fee",
        61: "Car Loan",
        62: "Student Loan",
        63: "Northwestern Mutual",
        64: "Tutoring",
        65: "Goddard",
        66: "Gilda",
        68: "Groceries",
    }
    for row, name in expense_map.items():
        val = ws[f"H{row}"].value
        if val and val > 0:
            expense_items.append({"name": name, "amount": round(val)})
    expense_items.sort(key=lambda x: x["amount"], reverse=True)
    monthly_expenses = sum(item["amount"] for item in expense_items)

    # ── Snapshot date ──
    date_cell = ws[f"{col}1"].value
    if isinstance(date_cell, datetime):
        snapshot_date = date_cell.strftime("%Y-%m-%d")
    elif isinstance(date_cell, (int, float)):
        from datetime import timedelta
        snapshot_date = (datetime(1899, 12, 30) + timedelta(days=int(date_cell))).strftime("%Y-%m-%d")
    else:
        snapshot_date = str(date_cell) if date_cell else "unknown"

    return {
        "snapshot_date": snapshot_date,
        "post_tax": {
            "total": post_tax_total,
            "accounts": post_tax_accounts,
        },
        "anthropic_equity": {
            "total": anthropic_total,
            "note": "Estimated equity value",
        },
        "retirement": {
            "total": retirement_total,
            "accounts": retirement_accounts,
        },
        "real_estate": {
            "total": re_total,
            "properties": properties,
        },
        "liabilities": {
            "total": liab_total,
            "items": liab_items,
        },
        "net_worth": net_worth,
        "monthly_income": monthly_income,
        "monthly_expenses": monthly_expenses,
        "monthly_gap": monthly_income - monthly_expenses,
        "income_breakdown": income_items,
        "expense_breakdown": expense_items,
    }


if __name__ == "__main__":
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        xlsx_path = Path.home() / "Downloads" / "Finances Snapshot.xlsx"

    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    data = parse_xlsx(xlsx_path)

    out_path = Path(__file__).resolve().parent / "finances.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    print(f"Written: {out_path}")
    print(f"  Net Worth:  ${data['net_worth']:,.0f}")
    print(f"  Assets:     Post-Tax ${data['post_tax']['total']:,.0f} + "
          f"Anthropic ${data['anthropic_equity']['total']:,.0f} + "
          f"Retirement ${data['retirement']['total']:,.0f} + "
          f"RE ${data['real_estate']['total']:,.0f}")
    print(f"  Liabilities: ${data['liabilities']['total']:,.0f}")
    print(f"  Cash Flow:  ${data['monthly_income']:,.0f} - "
          f"${data['monthly_expenses']:,.0f} = "
          f"${data['monthly_gap']:,.0f}/mo")
